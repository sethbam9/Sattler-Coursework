import sqlite3

# https://www.esv.org/resources/esv-global-study-bible/list-of-abbreviations/
bible_books = [
'Gen', 'Ex', 'Lev', 'Num', 'Deut', 'Josh', 'Judg', 'Ruth', '1 Sam', '2 Sam', '1 Kings',
'2 Kings', '1 Chron', '2 Chron', 'Ezra', 'Neh', 'Est', 'Job', 'Ps', 'Prov', 'Eccles',
'Song', 'Isa', 'Jer', 'Lam', 'Ezek', 'Dan', 'Hos', 'Joel', 'Amos', 'Obad', 'Jonah', 'Mic',
'Nah', 'Hab', 'Zeph', 'Hag', 'Zech', 'Mal', 'Matt', 'Mark', 'Luke', 'John', 'Acts', 'Rom',
'1 Cor', '2 Cor', 'Gal', 'Eph', 'Phil', 'Col', '1 Thess', '2 Thess', '1 Tim', '2 Tim',
'Titus', 'Philem', 'Heb', 'James', '1 Pet', '2 Pet', '1 John', '2 John', '3 John', 'Jude', 'Rev'
]

# INDEX INFO
"""
['id', 'b', 'c', 'v', 't']
[('bible_version_key',), ('cross_reference',), ('key_english',),
('t_asv',), ('t_bbe',), ('t_dby',), ('t_kjv',), ('t_web',), ('t_ylt',), ('t_wbt',),
('people',)]
"""

class Scripture:

    def __init__(self, translation, book, chapter):
        self.tran = translation
        self.bk = bible_books.index(book) + 1
        self.ch = chapter

    # Load the dataset and return as conn
    def loadData(self):
        db = r"C:\Users\sethb\OneDrive\Desktop\Sattler College\5. Fall 2020\CS 202 Object-Oriented Design\bible-sqlite.db"
        conn = sqlite3.connect(db)
        return conn

    # Return all table names from the dataset
    def dataContents(self):
        conn = self.loadData()
        cursorObj = conn.cursor()
        cursorObj.execute('SELECT name from sqlite_master where type= "table"')
        return cursorObj.fetchall()

    # Return all content names from a table
    def tableContents(self, table):
        conn = self.loadData()
        cursor = conn.execute("SELECT * FROM %s" % (table))
        names = list(map(lambda x: x[0], cursor.description))
        return names

    # Return the whole passage as a single string.
    def getPassage(self):
        conn = self.loadData()
        scripture = []
        chapter = conn.execute("SELECT * FROM %s WHERE b = %s AND c = %s"
            % (self.tran, self.bk, self.ch))
        for i in chapter:
            c = str(i[2])
            v = str(i[3])
            t = i[4]
            # t = '-\n'.join(t_temp[i:i+n] for i in range(0,len(t_temp),n))
            all = "%s:%s %s\n" % (c, v, t)
            scripture.append(all)
        passage = ''.join(scripture)
        return passage

    # Return the whole passage as a list of verses.
    def getVerses(self):
        conn = self.loadData()
        scripture = []
        chapter = conn.execute("SELECT * FROM %s WHERE b = %s AND c = %s"
            % (self.tran, self.bk, self.ch))
        for i in chapter:
            v = str(+i[3])
            t = i[4]
            all = "%s %s" % (v, t)
            scripture.append(all)
        return scripture

    # Return the whole passage as a list of verses.
    def getVerse(self):
        conn = self.loadData()
        scripture = []
        chapter = conn.execute("SELECT * FROM %s WHERE b = %s AND c = %s"
            % (self.tran, self.bk, self.ch))
        for i in chapter:
            v = str(+i[3])
            t = i[4]
            all = "%s %s" % (v, t)
            scripture.append(all)
        return scripture

scripture = Scripture('t_kjv', 'Ps', 16)
verses = scripture.getVerses()
for verse in verses:
    print(verse)


def Verses(ch, vs):
    global scripture
    conn = scripture.loadData()
    verse = []
    b = bible_books.index('Isa') + 1
    vs = conn.execute("SELECT * FROM t_kjv WHERE b = %s AND c = %s AND v = %s"
        % (b, ch, vs))
    for i in vs:
        verse.append(i[4])
    passage = ''.join(verse)
    return passage

from openpyxl import load_workbook

xfile = r"C:\Users\sethb\Downloads\Isaiah Notes.xlsx"

def getIt(x):
    list = []
    data = str(x)
    col = data.find(':')
    ch = int(data[2:col])
    vs = int(data[col+1:len(data)])
    list.append(ch)
    list.append(vs)
    return list

def makeIt(f):
    wb = load_workbook(f)
    sheet = wb.active
    for i in range(2, 69):
        l = getIt(sheet.cell(row=i, column=1).value)
        v = Verses(l[0], l[1])
        sheet.cell(row=i, column=6, value=v)
    wb.save(xfile)

# makeIt(xfile)
